import threading
import socket
import openpyxl
import subprocess
from tkinter import *
from tkinter import ttk, messagebox
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import platform

# Dicionário de unidades
subredes_unidades = {
    "Arouche": "10.1.0.0/24",
    "Hospital Santa Cecilia": "10.2.0.0/24",
    "Almoxarifado": "10.3.0.0/24",
    "Ressonancia": "10.4.0.0/24",
    "Heliópolis": "10.5.0.0/24",
    "Franco da Rocha": "10.6.0.0/24"
}

# Configurações para ocultar janela do cmd no Windows
subprocess_flags = {
    "startupinfo": None,
    "creationflags": 0
}

if platform.system() == "Windows":
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    subprocess_flags["startupinfo"] = si
    subprocess_flags["creationflags"] = subprocess.CREATE_NO_WINDOW

# Função para converter IP em número (para ordenação correta)
def ip_para_inteiro(ip):
    partes = ip.split('.')
    return int(partes[0]) * 256**3 + int(partes[1]) * 256**2 + int(partes[2]) * 256 + int(partes[3])

# Função para executar nmap de forma silenciosa
def executar_nmap(hosts, argumentos):
    comando = ["nmap"] + argumentos.split() + [hosts]
    resultado = subprocess.run(comando, capture_output=True, text=True, **subprocess_flags)
    return resultado.stdout

# Função principal que faz o inventário
def escanear_rede():
    nome_unidade = combo_rede.get()
    rede = subredes_unidades.get(nome_unidade)
    modo_agressivo = var_agressivo.get()

    if not rede:
        messagebox.showwarning("Atenção", "Escolha uma unidade válida")
        return

    botao_iniciar.config(state=DISABLED)
    label_status.config(text="Escaneando... Por favor, aguarde.")
    barra_progresso['value'] = 0

    def processo():
        dados = []

        argumentos = '-sn -R'
        if modo_agressivo:
            argumentos = '-p 135,139,445,3389,80,443 -T4'

        try:
            resultado = executar_nmap(rede, argumentos)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha na varredura:\n{str(e)}")
            botao_iniciar.config(state=NORMAL)
            return

        linhas = resultado.splitlines()
        hosts = []
        ip_atual = None

        for linha in linhas:
            if "Nmap scan report for" in linha:
                partes = linha.split(" ")
                ip_atual = partes[-1]
                hosts.append(ip_atual)

        total = len(hosts)

        for i, host in enumerate(hosts):
            try:
                nome_host = ""
                try:
                    nome_host = socket.gethostbyaddr(host)[0]
                except:
                    nome_host = "Não identificado"

                mac = "Desconhecido"
                fabricante = "Desconhecido"

                try:
                    resultado_os = executar_nmap(host, "-O")
                    for linha_os in resultado_os.splitlines():
                        if "OS details:" in linha_os:
                            sistema_operacional = linha_os.split(":", 1)[1].strip()
                            break
                    else:
                        sistema_operacional = "Não identificado"
                except:
                    sistema_operacional = "Não identificado"

                dados.append({
                    "IP": host,
                    "IP_Ordem": ip_para_inteiro(host),
                    "Nome do Host": nome_host,
                    "MAC": mac,
                    "Fabricante": fabricante,
                    "Sistema Operacional": sistema_operacional
                })
            except:
                continue

            barra_progresso['value'] = (i + 1) / total * 100
            janela.update_idletasks()

        dados.sort(key=lambda x: x["IP_Ordem"])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventário da Rede"

        cabecalhos = ["IP", "Nome do Host", "MAC", "Fabricante", "Sistema Operacional"]
        ws.append(cabecalhos)

        for col in range(1, len(cabecalhos)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for item in dados:
            ws.append([
                item["IP"],
                item["Nome do Host"],
                item["MAC"],
                item["Fabricante"],
                item["Sistema Operacional"]
            ])

        for coluna in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in coluna)
            col_letter = get_column_letter(coluna[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(dados)+1}"

        data_hora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"inventario_{nome_unidade.replace(' ', '_')}_{data_hora}.xlsx"
        wb.save(nome_arquivo)

        label_status.config(text=f"Escaneamento concluído! Arquivo salvo: {nome_arquivo}")
        botao_iniciar.config(state=NORMAL)

    threading.Thread(target=processo).start()

# Cria a janela principal
janela = Tk()
janela.title("Inventário de Rede")
janela.geometry("450x280")

Label(janela, text="Escolha a unidade para escanear:").pack(pady=5)
combo_rede = ttk.Combobox(janela, values=list(subredes_unidades.keys()), width=30)
combo_rede.pack()
combo_rede.set("Arouche")

var_agressivo = BooleanVar()
Checkbutton(janela, text="Modo agressivo (tenta detectar mais)", variable=var_agressivo).pack(pady=5)

botao_iniciar = Button(janela, text="Iniciar Escaneamento", command=escanear_rede)
botao_iniciar.pack(pady=10)

barra_progresso = ttk.Progressbar(janela, length=300, mode='determinate')
barra_progresso.pack(pady=5)

label_status = Label(janela, text="")
label_status.pack(pady=5)

janela.mainloop()
